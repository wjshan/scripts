from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell, get_type
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.styles import (
    Fill, Font, PatternFill, Alignment, Border, Side, Color,
    PatternFill, GradientFill, Fill, Font, DEFAULT_FONT, NumberFormatDescriptor,
    Protection, NamedStyle
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import get_column_interval, get_column_letter
from lxml import etree
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from dateutil.parser import parse

type_map = {
    'n': float,
    's': str,
    'b': bool,
    'd': parse
}


def xml2excel(xml_string):
    root = etree.XML(xml_string)
    work_book = Workbook()
    for index, work_sheet_tag in enumerate(root.xpath('WorkSheet')):
        title = work_sheet_tag.get('title', None)
        if index == 0:
            work_sheet = work_book.active
            work_sheet.title = title
        else:
            work_sheet = work_book.create_sheet(title)
        cells = work_sheet_tag.xpath('Cell')
        cols = work_sheet_tag.xpath('cols/col')
        rows = work_sheet_tag.xpath('rows/row')
        alignments = work_sheet_tag.xpath('alignments/alignment')
        fills = work_sheet_tag.xpath('fills/fill')
        fonts = work_sheet_tag.xpath('fonts/font')
        borders = work_sheet_tag.xpath('borders/border')
        protections = work_sheet_tag.xpath('protections/protection')
        data_validation_list = work_sheet_tag.xpath('DataValidationList')
        for col in cols:
            min_col = int(col.get('min'))
            max_col = int(col.get('max'))
            work_sheet.column_dimensions[get_column_letter(min_col)] = ColumnDimension(work_sheet,
                                                                                       **dict(col.attrib, style=None))
        for row in rows:
            index = row.attrib.pop('index')
            work_sheet.row_dimensions[int(index)] = RowDimension(work_sheet, **dict(row.attrib, s=None))
        style_map = {}
        for alignment in alignments:
            _id = alignment.attrib.pop('id')
            style_map[_id] = Alignment.from_tree(alignment)
        for fill in fills:
            _id = fill.attrib.pop('id')
            style_map[_id] = Fill.from_tree(fill)
        for font in fonts:
            _id = font.attrib.pop('id')
            style_map[_id] = Font.from_tree(font)
        for protection in protections:
            _id = protection.attrib.pop('id')
            style_map[_id] = Protection.from_tree(protection)
        for border in borders:
            _id = border.attrib.pop('id')
            style_map[_id] = Border.from_tree(border)
        for cell in cells:
            attrs = cell.attrib
            date_type = attrs.get('type')
            if date_type:
                value = type_map[date_type](attrs.get('value', None))
            else:
                value = attrs.get('value', None)
            cell = work_sheet.cell(column=int(attrs['column']), row=int(attrs['row']), value=value)
            min_row = attrs.get('min_row')
            max_row = attrs.get('max_row')
            min_col = attrs.get('min_col')
            max_col = attrs.get('max_col')
            if all([min_row, max_row, min_col, max_col]):
                work_sheet.merge_cells(start_row=int(min_row), start_column=int(min_col), end_row=int(max_row),
                                       end_column=int(max_col))
            if "fill_id" in attrs:
                cell.fill = style_map[attrs['fill_id']]
            if "alignment_id" in attrs:
                cell.alignment = style_map[attrs['alignment_id']]
            if "font_id" in attrs:
                cell.font = style_map[attrs['font_id']]
            if "protection_id" in attrs:
                cell.protection = style_map[attrs['protection_id']]
            if "border_id" in attrs:
                cell.border = style_map[attrs['border_id']]
    return work_book


if __name__ == '__main__':
    with open('专业分包统计表.xml') as f:
        wb = xml2excel(f.read())
    wb.save('test.xlsx')
