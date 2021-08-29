# -*-coding:utf-8 -*-
# @author: DanDan
# @contact: 454273687@qq.com
# @file: sheet.py
# @time: 2021/8/29 15:40
# @desc:
from .serializer import SerializerAble
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from .cell import CellSerializer, RowCellsSerializer
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl import Workbook
from lxml import etree


class ColumnSerializer(object):
    tag_name = 'Column'

    def __init__(self, index, attrs):
        self.index = index
        self.attrs = attrs


class SheetSerializer(SerializerAble):
    tag_name = 'sheet'

    def __init__(self, title=None):
        self.title = title
        self.rows = []
        self.column_style = []
        self._name = id(self)

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, name):
        if name:
            self._name = name

    @classmethod
    def from_excel(cls, work_sheet: Worksheet, *args, **kwargs):
        self = cls(work_sheet.title)
        merged_cell_map = {x.start_cell.coordinate: (x.min_row, x.max_row, x.min_col, x.max_col) for x in
                           work_sheet.merged_cells}
        row_dimension_map = {}
        for row_index, row_dimension in work_sheet.row_dimensions.items():
            row_dimension_map[row_index] = row_dimension
        row_index = col_index = 1
        last_row_index = last_col_index = 1
        for column_dimension in work_sheet.column_dimensions.values():
            column_dimension.reindex()
            self.column_style.append(dict(column_dimension))
        for row in work_sheet.iter_rows():
            cells = []
            for cell in row:
                if not isinstance(cell, Cell):
                    col_index += 1
                    continue
                cell_coordinate = cell.coordinate
                row_span = col_span = 0
                if cell_coordinate in merged_cell_map:
                    min_row, max_row, min_col, max_col = merged_cell_map[cell_coordinate]
                    row_span = max_row - min_row
                    col_span = max_col - min_col
                skip_cols = col_index - last_col_index
                cells.append(
                    CellSerializer.from_excel(cell, row_span=row_span, col_span=col_span, skip_cols=skip_cols)
                )
                col_index += 1
                last_col_index = col_index
            if cells:
                skip_rows = row_index - last_row_index
                row_instance = row_dimension_map.get(row_index, None)
                row_serializer = RowCellsSerializer.from_excel(row_instance, index=row_index, cells=cells,
                                                               skip_rows=skip_rows)
                self.rows.append(row_serializer)
            row_index += 1
            if cells:
                last_row_index = row_index
            last_col_index = col_index = 1
        return self

    def to_excel(self, parent: Workbook, *args, **kwargs):
        index = kwargs.get('index')
        sheet = parent.create_sheet(self.title, index=index)
        row_index = 1
        for row in self.rows:
            skip_row = row.to_excel(sheet, row_index=row_index)
            row_index = row_index + skip_row + 1

    @classmethod
    def from_xml(cls, node, *args, **kwargs):
        title = cls.convert_python_value(node.attrib.get('title'), 'str')
        self = cls(title)
        self.name = self.convert_python_value(node.attrib.get('name'), 'int')
        style_nodes = kwargs.get('style_node') or node
        for style_node in style_nodes.xpath(f"//ColumnStyle[@name={self.name}]"):
            self.column_style.append(style_node.attrib)
        for row_node in node.xpath(f"{RowCellsSerializer.tag_name}"):
            self.rows.append(RowCellsSerializer.from_xml(row_node, style_node=style_nodes))
        return self

    def to_xml(self, *args, **kwargs):
        name = self.convert_xml_value(self.name)
        sheet_tag = etree.Element(self.tag_name, title=self.convert_xml_value(self.title), name=name)
        style_tags = []
        for column_attr in self.column_style:
            style_tags.append(
                etree.Element('ColumnStyle', name=name, **column_attr)
            )
        for row in self.rows:
            row_node, style_nodes = row.to_xml()
            sheet_tag.append(row_node)
            style_tags.extend(style_nodes)
        return sheet_tag, style_tags
