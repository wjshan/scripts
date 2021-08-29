import json

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell, get_type
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.styles import (
    Fill, Font, PatternFill, Alignment, Border, Side, Color,
    PatternFill, GradientFill, Fill, Font, DEFAULT_FONT, NumberFormatDescriptor,
    Protection, NamedStyle
)
from openpyxl.worksheet.datavalidation import DataValidation
from lxml import etree
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.compat import safe_string


def excel2xml(file: str):
    work_book = load_workbook(file)
    root_tag = etree.Element('WorkBook')
    for work_sheet in work_book:
        work_sheet_tag = etree.Element('WorkSheet', title=work_sheet.title)
        merged_cell_map = {x.start_cell.coordinate: (x.min_row, x.max_row, x.min_col, x.max_col) for x in
                           work_sheet.merged_cells}
        col_tag = work_sheet.column_dimensions.to_tree()
        work_sheet_tag.append(col_tag)
        rows_tag = etree.SubElement(work_sheet_tag, 'rows')
        for row_index, row_dimension in work_sheet.row_dimensions.items():
            etree.SubElement(rows_tag, 'row', **dict(row_dimension, index=safe_string(row_dimension.index)))
        data_validation_list_target = work_sheet.data_validations.to_tree('DataValidationList')
        work_sheet_tag.append(data_validation_list_target)
        style_fields = ['alignment', 'fill', 'font', 'protection', 'border']
        style_group_map = {x: etree.SubElement(work_sheet_tag, f"{x}s") for x in style_fields}
        style_attr_cache = []
        for row in work_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, Cell):
                    cell_tag = etree.Element('Cell', column=safe_string(cell.column),
                                             row=safe_string(cell.row), value=safe_string(cell.value),
                                             number_format=cell.number_format,
                                             type=get_type(type(cell.value), cell.value) or '')
                    if cell.has_style:
                        for style_name in style_fields:
                            style_attr = getattr(cell, style_name)
                            style_attrs = dict(style_attr._StyleProxy__target)
                            if not style_attrs:
                                continue
                            if style_attrs not in style_attr_cache:
                                style_id = len(style_attr_cache)
                                group_style_tag = style_group_map[style_name]
                                style_attr_cache.append(style_attrs)
                                style_tag = style_attr.to_tree(style_name)
                                style_tag.attrib['id'] = safe_string(style_id)
                                group_style_tag.append(style_tag)
                            else:
                                style_id = style_attr_cache.index(style_attrs)
                            cell_tag.attrib[f'{style_name}_id'] = safe_string(style_id)

                    if cell.coordinate in merged_cell_map:
                        min_row, max_row, min_col, max_col = merged_cell_map[cell.coordinate]
                        cell_tag.attrib.update({
                            "min_row": str(min_row),
                            "max_row": str(max_row),
                            "min_col": str(min_col),
                            "max_col": str(max_col),
                        })
                    if not (cell.value or cell.coordinate in merged_cell_map):
                        continue
                    work_sheet_tag.append(cell_tag)
            etree.SubElement(work_sheet_tag, 'br')
        root_tag.append(work_sheet_tag)
    return root_tag


if __name__ == '__main__':
    import sys

    excel_path, xml_path = sys.argv[1:]

    work_book_xml = excel2xml(excel_path)
    with open(xml_path, 'wb') as f:
        f.write(etree.tostring(work_book_xml, pretty_print=True, encoding='utf-8'))
